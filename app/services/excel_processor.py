import asyncio
import io
import logging
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)


@dataclass
class GenerationRequest:
    """Describe content that should be generated for an Excel row."""

    sheet_name: str
    row_idx: int
    cycle: str
    post_type: str
    topic: str
    generate_content: bool
    generate_hashtags: bool


@dataclass
class GenerationResult:
    """Represent generated content for an Excel row."""

    sheet_name: str
    row_idx: int
    content: str | None = None
    hashtags: str | None = None


class ExcelProcessorService:
    """
    Parse, process, and generate Excel files.

    File operations are performed in memory to keep the service stateless.
    """

    @staticmethod
    def _load_workbook_sync(file_bytes: bytes) -> Workbook:
        """
        Load an Excel workbook synchronously from in-memory bytes.

        This blocking operation is intended to run in a separate thread.
        """

        return load_workbook(filename=io.BytesIO(file_bytes))

    @staticmethod
    def _save_workbook_sync(workbook: Workbook) -> bytes:
        """Save an Excel workbook synchronously to in-memory bytes."""

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _collect_generation_requests_sync(
        workbook: Workbook,
    ) -> list[GenerationRequest]:
        """Collect eligible Excel rows that require generated content."""

        required_columns = {
            "Cykl",
            "Typ posta",
            "Temat",
            "Treść posta",
        }

        requests: list[GenerationRequest] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            logger.info("Inspecting sheet: %s", sheet_name)

            headers = {}

            for cell in sheet[1]:
                if isinstance(cell.value, str) and cell.value.strip():
                    headers[cell.value.strip()] = cell.column

            if not required_columns.issubset(headers):
                missing_columns = required_columns - headers.keys()

                logger.info(
                    "Skipping sheet %s: missing required columns: %s",
                    sheet_name,
                    sorted(missing_columns),
                )
                continue

            logger.info("Processing sheet: %s", sheet_name)

            for row_idx in range(2, sheet.max_row + 1):
                row_values = [cell.value for cell in sheet[row_idx]]

                if any(
                    isinstance(value, str) and value.strip().casefold() == "stories"
                    for value in row_values
                ):
                    logger.info(
                        "Stopping sheet %s at Stories section (row %s).",
                        sheet_name,
                        row_idx,
                    )
                    break

                cycle = sheet.cell(
                    row=row_idx,
                    column=headers["Cykl"],
                ).value

                post_type = sheet.cell(
                    row=row_idx,
                    column=headers["Typ posta"],
                ).value

                topic = sheet.cell(
                    row=row_idx,
                    column=headers["Temat"],
                ).value

                missing_fields = []

                for field_name, value in (
                    ("Cykl", cycle),
                    ("Typ posta", post_type),
                    ("Temat", topic),
                ):
                    if value is None or (isinstance(value, str) and not value.strip()):
                        missing_fields.append(field_name)

                if missing_fields:
                    logger.debug(
                        "Skipping row %s in sheet %s: missing fields: %s",
                        row_idx,
                        sheet_name,
                        missing_fields,
                    )
                    continue

                content_cell = sheet.cell(
                    row=row_idx,
                    column=headers["Treść posta"],
                )

                content_is_empty = content_cell.value is None or (
                    isinstance(content_cell.value, str)
                    and not content_cell.value.strip()
                )

                generate_content = (
                    not isinstance(content_cell, MergedCell) and content_is_empty
                )

                generate_hashtags = False

                if "Hasła" in headers:
                    hashtag_cell = sheet.cell(
                        row=row_idx,
                        column=headers["Hasła"],
                    )

                    hashtags_are_empty = hashtag_cell.value is None or (
                        isinstance(hashtag_cell.value, str)
                        and not hashtag_cell.value.strip()
                    )

                    generate_hashtags = (
                        not isinstance(hashtag_cell, MergedCell) and hashtags_are_empty
                    )

                if not generate_content and not generate_hashtags:
                    continue

                requests.append(
                    GenerationRequest(
                        sheet_name=sheet_name,
                        row_idx=row_idx,
                        cycle=str(cycle),
                        post_type=str(post_type),
                        topic=str(topic),
                        generate_content=generate_content,
                        generate_hashtags=generate_hashtags,
                    )
                )

        return requests

    @staticmethod
    def _apply_generation_results_sync(
        workbook: Workbook,
        results: list[GenerationResult],
    ) -> Workbook:
        """Write generated content back to the workbook."""

        for result in results:
            sheet = workbook[result.sheet_name]

            headers = {}

            for cell in sheet[1]:
                if isinstance(cell.value, str) and cell.value.strip():
                    headers[cell.value.strip()] = cell.column

            if result.content is not None:
                content_cell = sheet.cell(
                    row=result.row_idx,
                    column=headers["Treść posta"],
                )

                if not isinstance(content_cell, MergedCell):
                    content_cell.value = result.content

            if result.hashtags is not None and "Hasła" in headers:
                hashtag_cell = sheet.cell(
                    row=result.row_idx,
                    column=headers["Hasła"],
                )

                if not isinstance(hashtag_cell, MergedCell):
                    hashtag_cell.value = result.hashtags

        return workbook

    @staticmethod
    def _generate_mock_results(
        requests: list[GenerationRequest],
    ) -> list[GenerationResult]:
        """Generate temporary placeholder results for development."""

        results = []

        for request in requests:
            content = None
            hashtags = None

            if request.generate_content:
                content = (
                    "AI-GENERATED DRAFT. "
                    f"Cycle: {request.cycle}, "
                    f"post type: {request.post_type}, "
                    f"topic: {request.topic}"
                )

            if request.generate_hashtags:
                hashtags = "#ai #marketing #automation"

            results.append(
                GenerationResult(
                    sheet_name=request.sheet_name,
                    row_idx=request.row_idx,
                    content=content,
                    hashtags=hashtags,
                )
            )

        return results

    async def process_calendar_file(
        self,
        file_bytes: bytes,
    ) -> bytes:
        """Process the marketing calendar Excel file."""

        workbook = await asyncio.to_thread(
            self._load_workbook_sync,
            file_bytes,
        )

        requests = await asyncio.to_thread(
            self._collect_generation_requests_sync,
            workbook,
        )

        results = self._generate_mock_results(requests)

        processed_workbook = await asyncio.to_thread(
            self._apply_generation_results_sync,
            workbook,
            results,
        )

        return await asyncio.to_thread(
            self._save_workbook_sync,
            processed_workbook,
        )
