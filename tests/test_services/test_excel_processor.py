import io

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel_processor import ExcelProcessorService

HEADERS = [
    "Cykl",
    "Typ posta",
    "AI",
    "Temat",
    "Treść posta",
    "Hasła",
]


def create_workbook_bytes(
    rows: list[list[object]],
    headers: list[str] | None = None,
    sheet_name: str = "Test_Sheet",
) -> bytes:
    """Create an in-memory Excel workbook for tests."""

    workbook = Workbook()
    worksheet = workbook.active

    assert worksheet is not None

    worksheet.title = sheet_name
    worksheet.append(headers or HEADERS)

    for row in rows:
        worksheet.append(row)

    output = io.BytesIO()
    workbook.save(output)

    return output.getvalue()


def load_processed_sheet(
    processed_bytes: bytes,
    sheet_name: str = "Test_Sheet",
) -> Worksheet:
    """Load a worksheet from processed workbook bytes."""

    workbook = load_workbook(io.BytesIO(processed_bytes))
    return workbook[sheet_name]


@pytest.fixture
def service() -> ExcelProcessorService:
    """Provide an ExcelProcessorService instance."""

    return ExcelProcessorService()


@pytest.mark.asyncio
async def test_populates_missing_content_and_hashtags(
    service: ExcelProcessorService,
) -> None:
    """Populate missing content and hashtags for an eligible row."""

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q1",
                "Carousel",
                False,
                "Marketing automation",
                None,
                None,
            ]
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    content = worksheet.cell(row=2, column=5).value
    hashtags = worksheet.cell(row=2, column=6).value

    assert isinstance(content, str)
    assert "AI-GENERATED DRAFT" in content
    assert "Marketing automation" in content
    assert "Carousel" in content
    assert "Q1" in content

    assert hashtags == "#ai #marketing #automation"


@pytest.mark.asyncio
async def test_preserves_existing_content_and_populates_missing_hashtags(
    service: ExcelProcessorService,
) -> None:
    """Preserve existing content while populating missing hashtags."""

    existing_content = "Existing promotional copy."

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q2",
                "Static post",
                True,
                "Summer campaign",
                existing_content,
                None,
            ]
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    content = worksheet.cell(row=2, column=5).value
    hashtags = worksheet.cell(row=2, column=6).value

    assert content == existing_content
    assert hashtags == "#ai #marketing #automation"


@pytest.mark.asyncio
async def test_populates_missing_content_and_preserves_existing_hashtags(
    service: ExcelProcessorService,
) -> None:
    """Populate missing content while preserving existing hashtags."""

    existing_hashtags = "#brand #campaign"

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q3",
                "Reel",
                False,
                "Product launch",
                None,
                existing_hashtags,
            ]
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    content = worksheet.cell(row=2, column=5).value
    hashtags = worksheet.cell(row=2, column=6).value

    assert isinstance(content, str)
    assert "Product launch" in content
    assert hashtags == existing_hashtags


@pytest.mark.asyncio
async def test_preserves_existing_content_and_hashtags(
    service: ExcelProcessorService,
) -> None:
    """Do not overwrite existing content or hashtags."""

    existing_content = "Approved copy."
    existing_hashtags = "#approved #brand"

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q4",
                "Carousel",
                True,
                "Holiday promotion",
                existing_content,
                existing_hashtags,
            ]
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    assert worksheet.cell(row=2, column=5).value == existing_content
    assert worksheet.cell(row=2, column=6).value == existing_hashtags


@pytest.mark.asyncio
async def test_ignores_ai_column_value(
    service: ExcelProcessorService,
) -> None:
    """Process eligible rows regardless of the AI column value."""

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q1",
                "Static post",
                False,
                "Brand awareness",
                None,
                None,
            ],
            [
                "Q1",
                "Reel",
                True,
                "Customer story",
                None,
                None,
            ],
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    first_content = worksheet.cell(row=2, column=5).value
    second_content = worksheet.cell(row=3, column=5).value

    assert isinstance(first_content, str)
    assert isinstance(second_content, str)

    assert "Brand awareness" in first_content
    assert "Customer story" in second_content


@pytest.mark.asyncio
async def test_processes_sheet_without_ai_column(
    service: ExcelProcessorService,
) -> None:
    """Process eligible rows when the AI column is absent."""

    headers = [
        "Cykl",
        "Typ posta",
        "Temat",
        "Treść posta",
        "Hasła",
    ]

    file_bytes = create_workbook_bytes(
        headers=headers,
        rows=[
            [
                "Q2",
                "Carousel",
                "Lead generation",
                None,
                None,
            ]
        ],
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    content = worksheet.cell(row=2, column=4).value
    hashtags = worksheet.cell(row=2, column=5).value

    assert isinstance(content, str)
    assert "Lead generation" in content
    assert hashtags == "#ai #marketing #automation"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("cycle", "post_type", "topic"),
    [
        (None, "Carousel", "Campaign topic"),
        ("", "Carousel", "Campaign topic"),
        ("   ", "Carousel", "Campaign topic"),
        ("Q1", None, "Campaign topic"),
        ("Q1", "", "Campaign topic"),
        ("Q1", "   ", "Campaign topic"),
        ("Q1", "Carousel", None),
        ("Q1", "Carousel", ""),
        ("Q1", "Carousel", "   "),
    ],
)
async def test_skips_rows_with_missing_required_values(
    service: ExcelProcessorService,
    cycle: object,
    post_type: object,
    topic: object,
) -> None:
    """Skip rows missing cycle, post type, or topic."""

    file_bytes = create_workbook_bytes(
        rows=[
            [
                cycle,
                post_type,
                False,
                topic,
                None,
                None,
            ]
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    assert worksheet.cell(row=2, column=5).value is None
    assert worksheet.cell(row=2, column=6).value is None


@pytest.mark.asyncio
async def test_treats_whitespace_content_and_hashtags_as_empty(
    service: ExcelProcessorService,
) -> None:
    """Treat whitespace-only output cells as empty."""

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q1",
                "Reel",
                False,
                "New collection",
                "   ",
                "   ",
            ]
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    content = worksheet.cell(row=2, column=5).value
    hashtags = worksheet.cell(row=2, column=6).value

    assert isinstance(content, str)
    assert "New collection" in content
    assert hashtags == "#ai #marketing #automation"


@pytest.mark.asyncio
async def test_stops_processing_at_stories_section(
    service: ExcelProcessorService,
) -> None:
    """Stop processing rows when the Stories section begins."""

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q1",
                "Carousel",
                False,
                "Main feed post",
                None,
                None,
            ],
            [
                "STORIES",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "Q1",
                "Story",
                False,
                "Story content",
                None,
                None,
            ],
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    feed_content = worksheet.cell(row=2, column=5).value
    story_content = worksheet.cell(row=4, column=5).value

    assert isinstance(feed_content, str)
    assert "Main feed post" in feed_content

    assert story_content is None


@pytest.mark.asyncio
async def test_detects_stories_marker_in_any_column(
    service: ExcelProcessorService,
) -> None:
    """Stop processing when the Stories marker appears in any column."""

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q1",
                "Carousel",
                False,
                "Feed campaign",
                None,
                None,
            ],
            [
                None,
                None,
                None,
                " STORIES ",
                None,
                None,
            ],
            [
                "Q1",
                "Reel",
                False,
                "Should not be processed",
                None,
                None,
            ],
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    assert isinstance(
        worksheet.cell(row=2, column=5).value,
        str,
    )

    assert worksheet.cell(row=4, column=5).value is None


@pytest.mark.asyncio
async def test_skips_sheet_with_missing_required_columns(
    service: ExcelProcessorService,
) -> None:
    """Leave sheets unchanged when required columns are missing."""

    file_bytes = create_workbook_bytes(
        headers=[
            "Cykl",
            "Typ posta",
            "Temat",
            "Hasła",
        ],
        rows=[
            [
                "Q1",
                "Carousel",
                "Campaign without content column",
                None,
            ]
        ],
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    assert worksheet.cell(row=2, column=1).value == "Q1"
    assert worksheet.cell(row=2, column=2).value == "Carousel"
    assert worksheet.cell(row=2, column=3).value == "Campaign without content column"


@pytest.mark.asyncio
async def test_processes_multiple_eligible_rows(
    service: ExcelProcessorService,
) -> None:
    """Process multiple eligible rows in the same worksheet."""

    file_bytes = create_workbook_bytes(
        rows=[
            [
                "Q1",
                "Carousel",
                False,
                "Campaign one",
                None,
                None,
            ],
            [
                "Q1",
                "Reel",
                False,
                "Campaign two",
                None,
                None,
            ],
            [
                "Q2",
                "Static post",
                False,
                "Campaign three",
                None,
                None,
            ],
        ]
    )

    processed_bytes = await service.process_calendar_file(file_bytes)
    worksheet = load_processed_sheet(processed_bytes)

    for row_idx in range(2, 5):
        content = worksheet.cell(row=row_idx, column=5).value
        hashtags = worksheet.cell(row=row_idx, column=6).value

        assert isinstance(content, str)
        assert hashtags == "#ai #marketing #automation"


@pytest.mark.asyncio
async def test_processes_only_compatible_sheets(
    service: ExcelProcessorService,
) -> None:
    """Process compatible sheets and leave unrelated sheets untouched."""

    workbook = Workbook()

    compatible_sheet = workbook.active
    assert compatible_sheet is not None

    compatible_sheet.title = "Instagram"
    compatible_sheet.append(HEADERS)
    compatible_sheet.append(
        [
            "Q1",
            "Carousel",
            False,
            "Instagram campaign",
            None,
            None,
        ]
    )

    unrelated_sheet = workbook.create_sheet("Stories")
    unrelated_sheet.append(
        [
            "Date",
            "Format",
            "Story text",
        ]
    )
    unrelated_sheet.append(
        [
            "2026-08-01",
            "Story",
            "Existing story",
        ]
    )

    output = io.BytesIO()
    workbook.save(output)

    processed_bytes = await service.process_calendar_file(output.getvalue())

    processed_workbook = load_workbook(io.BytesIO(processed_bytes))

    instagram_content = (
        processed_workbook["Instagram"]
        .cell(
            row=2,
            column=5,
        )
        .value
    )

    story_content = (
        processed_workbook["Stories"]
        .cell(
            row=2,
            column=3,
        )
        .value
    )

    assert isinstance(instagram_content, str)
    assert "Instagram campaign" in instagram_content

    assert story_content == "Existing story"
